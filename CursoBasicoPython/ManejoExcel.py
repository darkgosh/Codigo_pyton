import os
from openpyxl import Workbook
from openpyxl import load_workbook


"""
----------------------------------------------------------------------------------
                                  Codigo probado
---------------------------------------------------------------------------------
"""
#Escribir sobre las celdas
# wb = Workbook() # busca el archivo un nivel arriva del direcotrio donde corre el punto Py
# ws = wb.active
# wb.save("newfile.xlsx")
#ws2['E7']= "Test"
#wb2.save("variables.xlsx")


wb2= load_workbook("variables.xlsx")
ws2 = wb2['Hoja1']

ApiKey = ws2["C2"]
Secret = ws2["C3"]
print(ApiKey.value , Secret.value)


"""
----------------------------------------------------------------------------------
                                   AREA DE PRUEBA
---------------------------------------------------------------------------------
"""




"""
----------------------------------------------------------------------------------
                                    AREA DE PRUEBA
---------------------------------------------------------------------------------
"""

miarchivo = r'G:\\My Drive\\Projectos de Software\\PROYECTOS\\LaloBot\\PC_WINDOWS\\LaloBot\\Variables.xlsx'
# # leer el archivo
book = openpyxl.load_workbook(miarchivo,data_only=True)
print(book.sheetnames)

hoja1 = pd.read_excel(io = miarchivo, 
                      sheet_name= "Hoja1",
                      header= 0, 
                      index_col= None,
                      usecols= "A:E", 
                      engine="openpyxl"
                      )
#print(hoja1.head(1))
# renglones
ApiKey = (hoja1.iloc[[0],[2]])
print(ApiKey)

# Secret = str(sheet.cell_value(3, 3))
"""